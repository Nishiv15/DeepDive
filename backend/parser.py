import io
import fitz
from docx import Document as DocxDocument
from openpyxl import load_workbook


# Text Extraction

def extract_text_from_pdf(file_bytes: bytes) -> list[dict]:
    """Returns list of {page, text} dicts."""
    doc = fitz.open(stream=file_bytes, filetype="pdf")
    pages = []
    for i, page in enumerate(doc):
        text = page.get_text("text").strip()
        if text:
            pages.append({"page": i + 1, "text": text})
    return pages


def extract_text_from_docx(file_bytes: bytes) -> list[dict]:
    doc = DocxDocument(io.BytesIO(file_bytes))
    full_text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    return [{"page": 1, "text": full_text}]


def extract_text_from_txt(file_bytes: bytes) -> list[dict]:
    text = file_bytes.decode("utf-8", errors="ignore")
    return [{"page": 1, "text": text}]


def extract_text_from_xlsx(file_bytes: bytes) -> list[dict]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    rows = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = "\t".join(str(c) for c in row if c is not None)
            if row_text.strip():
                rows.append(row_text)
    return [{"page": 1, "text": "\n".join(rows)}]


EXTRACTORS = {
    "application/pdf": extract_text_from_pdf,
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": extract_text_from_docx,
    "text/plain": extract_text_from_txt,
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": extract_text_from_xlsx,
}


def extract_text(file_bytes: bytes, content_type: str) -> list[dict]:
    extractor = EXTRACTORS.get(content_type)
    if not extractor:
        raise ValueError(f"Unsupported file type: {content_type}")
    return extractor(file_bytes)


# Chunking

def chunk_pages(pages: list[dict], chunk_size: int, chunk_overlap: int) -> list[dict]:
    """Splits page text into overlapping chunks, preserving page number."""
    chunks = []
    for page_data in pages:
        text = page_data["text"]
        page = page_data["page"]
        start = 0
        while start < len(text):
            end = start + chunk_size
            chunk_text = text[start:end].strip()
            if chunk_text:
                chunks.append({"page": page, "text": chunk_text, "start_char": start})
            start += chunk_size - chunk_overlap
    return chunks
